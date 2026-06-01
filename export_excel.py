# -*- coding: utf-8 -*-  # define a codificação do ficheiro para suportar acentos

from datetime import date, timedelta  # importa tipos de data e diferença de dias
import calendar                       # importa calendar para descobrir quantas semanas tem o mês
import re                             # importa re para tratar texto de "VAGA ABERTA"

from openpyxl import Workbook                                      # importa Workbook para criar ficheiros Excel
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side  # importa estilos de célula
from openpyxl.utils import get_column_letter                       # importa função para converter índice em letra de coluna

from engine import gerar_escala_semana, DIAS_LONGOS, semana_datas  # importa funções e constantes do motor de escala


# =========================
# CONSTANTES / CORES DO EXCEL
# =========================

SAIDA_XLSX = "Escala_Mensal.xlsx"  # nome do ficheiro Excel de saída (mensal)

FILL_MARGIN_RED = PatternFill("solid", fgColor="B00020")  # preenchimento vermelho para margens e separadores
FILL_TURNO_AM  = PatternFill("solid", fgColor="123B47")   # preenchimento azul escuro para turno da manhã
FILL_TURNO_PM  = PatternFill("solid", fgColor="1F4E5F")   # preenchimento azul escuro para turno da tarde/noite
FILL_VAGA      = PatternFill("solid", fgColor="B7B7B7")   # preenchimento cinza para vagas abertas

THIN_BORDER = Border(                                     # borda fina em todos os lados para células padrão
    left=Side(style="thin"),                              # borda esquerda fina
    right=Side(style="thin"),                             # borda direita fina
    top=Side(style="thin"),                               # borda superior fina
    bottom=Side(style="thin"),                            # borda inferior fina
)

DIAGONAL_BORDER = Border(                                 # borda com risco diagonal (para dias fora do mês)
    left=Side(style="thin"),                              # borda esquerda fina
    right=Side(style="thin"),                             # borda direita fina
    top=Side(style="thin"),                               # borda superior fina
    bottom=Side(style="thin"),                            # borda inferior fina
    diagonal=Side(style="thin"),                          # estilo da diagonal
    diagonalDown=True,                                    # ativa diagonal de cima para baixo
    diagonalUp=True,                                      # ativa diagonal de baixo para cima
)

CENTER = Alignment(                                       # alinhamento padrão das células
    horizontal="center",                                  # centra horizontalmente
    vertical="center",                                    # centra verticalmente
    wrap_text=True,                                       # permite quebra de linha dentro da célula
)


# nomes dos meses em português (para usar no cabeçalho da semana)
MESES_PT = [                                              # lista com o nome dos meses por extenso
    "janeiro", "fevereiro", "março", "abril",             # meses 1 a 4
    "maio", "junho", "julho", "agosto",                   # meses 5 a 8
    "setembro", "outubro", "novembro", "dezembro",        # meses 9 a 12
]

def _mes_pt(d):                                           # função para obter nome do mês em português
    """Devolve o nome do mês em português para uma data 'd'."""  # docstring descritiva
    return MESES_PT[d.month - 1]                          # devolve o nome do mês com base no índice (month começa em 1)


# =========================
# CONSTRUÇÃO DE CORES POR FUNCIONÁRIO (PARA EXCEL)
# =========================

def build_employee_colors(estado):                        # função que atribui cor a cada colaborador
    """
    Cria um dicionário emp_id -> PatternFill, atribuindo cores suaves
    e diferentes a cada funcionário.
    """                                                    # docstring da função
    base_colors = [                                       # lista de cores hexadecimais base
        "BDD7EE", "C6E0B4", "F8CBAD", "FFF2CC",           # bloco 1 de cores
        "D9E1F2", "F2DCDB", "E2EFDA", "E4DFEC",           # bloco 2 de cores
        "DDEBF7", "FFE699", "D6DCE5", "C9C9FF",           # bloco 3 de cores
    ]
    fills = {}                                            # dicionário resultado emp_id -> PatternFill
    employees = list(estado.get("employees", {}).keys())  # obtém lista de IDs de colaboradores do estado
    for idx, emp_id in enumerate(employees):              # percorre cada colaborador com índice
        color = base_colors[idx % len(base_colors)]       # escolhe cor ciclicamente na lista base_colors
        fills[emp_id] = PatternFill("solid", fgColor=color)  # cria PatternFill sólido com essa cor
    return fills                                          # devolve o dicionário de fills


def _turno_label_with_breaks(turno_str: str) -> str:      # formata texto do turno em três linhas
    """Formata o turno como três linhas (ex.: '07:00\\n-\\n11:00')."""  # docstring
    if "–" in turno_str:                                  # se o turno usa travessão
        ini, fim = turno_str.split("–", 1)                # separa início e fim pelo travessão
    else:                                                 # caso use hífen
        ini, fim = turno_str.split("-", 1)                # separa início e fim pelo hífen
    ini = ini.strip()                                     # remove espaços do início
    fim = fim.strip()                                     # remove espaços do fim
    return f"{ini}\n-\n{fim}"                             # devolve string em três linhas


def _turno_color(turno_str: str):                         # devolve a cor de fundo do bloco do turno
    """Devolve a cor do bloco do turno (manhã ou tarde/noite)."""  # docstring
    if "–" in turno_str:                                  # se usar travessão
        ini, _ = turno_str.split("–", 1)                  # obtém a parte inicial
    else:                                                 # se usar hífen
        ini, _ = turno_str.split("-", 1)                  # obtém a parte inicial
    hora = int(ini.strip().split(":")[0])                 # extrai a hora (HH) como inteiro
    if hora < 12:                                         # se hora for antes de meio-dia
        return FILL_TURNO_AM                              # usa cor de turno da manhã
    return FILL_TURNO_PM                                  # senão, cor de turno tarde/noite


# =========================
# DESENHO DA SEMANA NO EXCEL
# =========================

def escrever_tabela_semana(                               # função que escreve UMA semana na planilha
    ws,                                                   # worksheet do openpyxl onde vamos escrever
    start_row,                                            # linha inicial para desenhar esta semana
    segunda_semana,                                       # date da segunda-feira da semana
    semana_idx,                                           # índice da semana (1..4 ou 5) – só para título
    esc,                                                  # dicionário escala {date: {turno: [alocações]}}
    estado,                                               # estado completo da empresa
    base_col=2,                                           # coluna de início (2 = coluna B)
    mes_restrito=None,                                    # número do mês alvo (1..12) para marcar dias fora do mês
):
    """
    Escreve UMA semana na planilha Excel, a partir de 'start_row',
    usando o layout com coluna 'Janela (Turno)' à esquerda, blocos
    de turnos com MESMA ALTURA, margens vermelhas e texto centralizado.
    """                                                    # docstring da função

    lin = start_row                                       # ponteiro de linha atual
    e = estado                                            # atalho para o estado (configuração)
    emp_fills = build_employee_colors(e)                  # dicionário emp_id -> PatternFill baseado no estado

    def _fill_row(row, fill):                             # função interna para pintar uma linha inteira
        """Pinta a linha inteira (A até I) com uma cor de fundo."""  # docstring
        for c in range(1, 10):                            # percorre colunas de A (1) até I (9)
            ws.cell(row=row, column=c).fill = fill        # aplica o PatternFill passado na célula

    datas_semana = semana_datas(segunda_semana)           # gera lista de 7 datas (segunda..domingo) da semana

    # ==== margem superior da semana ====
    _fill_row(lin, FILL_MARGIN_RED)                       # pinta a primeira linha da semana de vermelho
    lin += 1                                              # avança para a linha seguinte

    # ==== título da semana ====
    ini_txt = datas_semana[0].strftime("%d/%m/%Y")        # texto da data inicial (segunda) da semana
    fim_txt = datas_semana[-1].strftime("%d/%m/%Y")       # texto da data final (domingo) da semana

    ws.merge_cells(start_row=lin, start_column=2, end_row=lin, end_column=9)  # junta as colunas B..I na linha do título
    t = ws.cell(row=lin, column=2, value=f"Semana {semana_idx} — {ini_txt} a {fim_txt}")  # cria célula com texto do título
    t.font = Font(bold=True, size=12)                     # deixa o título em negrito e tamanho 12
    t.alignment = CENTER                                  # centraliza o título
    ws.cell(row=lin, column=1).fill = FILL_MARGIN_RED     # mantém coluna A como margem vermelha
    lin += 1                                              # avança para a próxima linha

    # ==== cabeçalho: Janela (Turno) + dias ====
    ws.cell(row=lin, column=1).fill = FILL_MARGIN_RED     # coluna A do cabeçalho também vermelha

    h_turno = ws.cell(row=lin, column=2, value="Janela\n(Turno)")  # escreve label "Janela (Turno)" na coluna B
    h_turno.font = Font(bold=True)                        # cabeçalho em negrito
    h_turno.alignment = CENTER                            # centralizado
    h_turno.border = THIN_BORDER                          # borda fina à volta

    # dias da semana (colunas C..I)
    for idx, d in enumerate(datas_semana):                # percorre cada dia da semana com índice 0..6
        if mes_restrito is not None and d.month != mes_restrito:   # se temos mês alvo e este dia é de outro mês
            texto = DIAS_LONGOS[idx]                      # mostra apenas o nome do dia (Sem data)
        else:                                             # se o dia pertence ao mês alvo
            texto = f"{DIAS_LONGOS[idx]}\n{d.day} {_mes_pt(d)}"  # mostra nome do dia + número + mês por extenso

        c = ws.cell(row=lin, column=3 + idx, value=texto) # escreve o texto na coluna adequada (C..I)
        c.font = Font(bold=True)                          # cabeçalho em negrito
        c.alignment = CENTER                              # centralizado

        if mes_restrito is not None and d.month != mes_restrito:  # se dia está fora do mês alvo
            c.border = DIAGONAL_BORDER                    # usa borda com diagonal
        else:                                             # se dia é do mês alvo
            c.border = THIN_BORDER                        # borda fina normal

    lin += 1                                              # avança para a linha seguinte (corpo da semana)

    # ==== descobrir TODOS os turnos da semana ====
    turnos = set()                                        # conjunto para guardar todos os turnos únicos
    for d in datas_semana:                                # percorre cada dia da semana
        for turno in esc[d].keys():                       # percorre cada turno registado na escala para esse dia
            turnos.add(turno)                             # adiciona o turno ao conjunto (sem repetições)

    turnos = sorted(list(turnos), key=lambda t: t.split("–")[0])  # ordena turnos pela hora de início

    # ==== calcular ALTURA GLOBAL (slots) p/ TODOS os turnos ====
    slots_por_turno = 1                                   # define quantidade mínima de linhas por turno
    for d in datas_semana:                                # percorre cada dia da semana
        for turno in turnos:                              # percorre cada turno
            alocs = esc[d].get(turno, [])                 # obtém lista de alocações para este dia/turno
            n = len(alocs)                                # número de alocações
            if n > slots_por_turno:                       # se este turno precisa de mais linhas que o atual
                slots_por_turno = n                       # atualiza o máximo de slots

    if slots_por_turno < 1:                               # se por algum motivo ficou menor que 1
        slots_por_turno = 1                               # garante pelo menos 1 linha por turno

    # ==== desenhar bloco para cada turno ====
    for turno in turnos:                                  # percorre cada turno a desenhar
        top = lin                                         # primeira linha do bloco deste turno
        bottom = lin + slots_por_turno + 1                # última linha útil do bloco (+1 de respiro)

        # coluna A (margem vermelha) em todo o bloco deste turno
        for r in range(top, bottom + 1):                  # percorre todas as linhas do bloco
            ws.cell(row=r, column=1).fill = FILL_MARGIN_RED  # pinta coluna A de vermelho

        # rótulo do turno na coluna B, ocupando o bloco vertical todo
        ws.merge_cells(start_row=top, start_column=2, end_row=bottom, end_column=2)  # mescla células de B (turno)
        lab = ws.cell(row=top, column=2, value=_turno_label_with_breaks(turno))      # escreve o rótulo formatado
        lab.font = Font(bold=True, color="FFFFFF")        # texto em negrito, branco
        lab.fill = _turno_color(turno)                    # cor de fundo conforme o turno (AM/PM)
        lab.alignment = CENTER                            # centralizado
        lab.border = THIN_BORDER                          # borda fina

        # dicionário dia -> lista de alocações para este turno
        daily_items = {d: esc[d].get(turno, []) for d in datas_semana}  # mapeia cada dia às alocações desse turno

        row_ptr = top + 1                                 # primeira linha útil logo abaixo do rótulo

        # percorre as "linhas/slots" desse turno
        for _ in range(slots_por_turno):                  # para cada linha de slot
            for offset, d in enumerate(datas_semana):     # percorre os 7 dias da semana
                col = 3 + offset                          # coluna correspondente (C..I)
                c = ws.cell(row=row_ptr, column=col)      # obtém a célula
                c.alignment = CENTER                      # centraliza o texto

                if mes_restrito is not None and d.month != mes_restrito:  # se dia está fora do mês alvo
                    c.border = DIAGONAL_BORDER            # borda com diagonal
                    continue                              # não escreve texto nem fill

                c.border = THIN_BORDER                    # se é dia do mês alvo, borda fina normal

                itens = daily_items[d]                    # lista de itens (alocações) para este dia/turno
                texto = ""                                # texto da célula (inicialmente vazio)
                fill = None                               # fill da célula (inicialmente sem cor)

                if _ < len(itens):                        # se existe um item para este "slot"
                    it = itens[_]                         # obtém o item da posição atual
                    if isinstance(it, tuple):             # se o item é uma tupla (emp_id, role)
                        emp_id, role = it                 # descompacta emp_id e role
                        nome = e["employees"].get(emp_id, {}).get("nome", emp_id)  # obtém nome do colaborador
                        texto = f"{nome}\n({role.title()})"  # texto com nome e função em baixo
                        fill = emp_fills.get(emp_id)      # cor associada a este colaborador
                    else:                                 # caso seja string (ex.: "VAGA ABERTA (role)")
                        if isinstance(it, str) and it.strip().upper().startswith("VAGA ABERTA"):  # se começa com VAGA ABERTA
                            m = re.search(r"\((.*?)\)", it)  # tenta extrair função entre parênteses
                            funcao = m.group(1).strip().title() if m else ""  # obtém função legível se existir
                            if funcao:
                                texto = f"VAGA ABERTA\n({funcao})"  # texto com função
                            else:
                                texto = "VAGA ABERTA"       # texto genérico
                        else:
                            texto = str(it)                 # qualquer outra string, só converte para string
                        fill = FILL_VAGA                    # cor padrão para vagas

                c.value = texto                            # escreve o texto na célula
                if fill:
                    c.fill = fill                          # aplica o fill se existir

            row_ptr += 1                                   # avança para a próxima linha de slot

        # linha vermelha separando este turno do próximo
        _fill_row(bottom + 1, FILL_MARGIN_RED)             # pinta a linha imediatamente abaixo do bloco com vermelho
        lin = bottom + 2                                   # define nova linha corrente duas linhas abaixo

    _fill_row(lin, FILL_MARGIN_RED)                        # pinta margem inferior da semana
    lin += 1                                               # avança uma linha para depois da semana

    return lin                                             # devolve a próxima linha livre depois desta semana


# =========================
# FUNÇÃO PÚBLICA – GERAR EXCEL DO MÊS
# =========================

def gerar_e_exportar_excel_4_semanas(                      # função pública usada pelo app principal
    estado,                                                # estado completo da empresa
    primeira_segunda,                                      # data da primeira segunda-feira do calendário do mês
    ano_alvo,                                              # ano alvo (int)
    mes_alvo,                                              # mês alvo (1..12)
):
    """
    Gera a escala para o MÊS INTEIRO que contém 'primeira_segunda',
    desenhando tantas semanas (segunda–domingo) quanto forem necessárias,
    numa única folha Excel, mantendo o layout atual.
    """                                                    # docstring descritiva

    wb = Workbook()                                        # cria um novo Workbook Excel
    ws = wb.active                                         # obtém a folha ativa
    ws.title = "Escala Mensal"                             # define o título da folha

    col_widths = {                                         # dicionário com larguras de colunas
        1: 4,                                              # coluna A – margem
        2: 16,                                             # coluna B – "Janela (Turno)"
        3: 22,                                             # colunas C..I – dias da semana
        4: 22,
        5: 22,
        6: 22,
        7: 22,
        8: 22,
        9: 22,
    }
    for col, w in col_widths.items():                      # percorre cada coluna e largura
        ws.column_dimensions[get_column_letter(col)].width = w  # aplica a largura na coluna correspondente

    current_row = 2                                        # linha inicial onde a primeira semana será escrita

    primeiro_dia_mes = date(ano_alvo, mes_alvo, 1)         # constrói date para o primeiro dia do mês
    _, last_day_num = calendar.monthrange(ano_alvo, mes_alvo)  # obtém o último dia do mês (número)
    ultimo_dia_mes = date(ano_alvo, mes_alvo, last_day_num)    # constrói date para o último dia do mês

    offset_inicio = primeiro_dia_mes.weekday()             # índice do dia da semana do primeiro dia (0=Segunda)
    primeira_segunda_mes = primeiro_dia_mes - timedelta(days=offset_inicio)  # data da segunda antes/igual ao dia 1

    offset_fim = 6 - ultimo_dia_mes.weekday()              # diferença até chegar ao domingo
    ultimo_domingo_mes = ultimo_dia_mes + timedelta(days=offset_fim)  # data do último domingo exibido no calendário

    num_dias = (ultimo_domingo_mes - primeira_segunda_mes).days + 1  # número total de dias exibidos
    num_semanas = num_dias // 7                                   # quantidade de semanas (4 ou 5 normalmente)

    for semana_idx in range(num_semanas):               # percorre cada semana (0..num_semanas-1)
        segunda_semana = primeira_segunda_mes + timedelta(days=7 * semana_idx)  # data da segunda-feira da semana
        escala_semana = gerar_escala_semana(estado, segunda_semana)             # gera escala daquela semana

        current_row = escrever_tabela_semana(           # escreve o bloco desta semana na planilha
            ws,                                         # worksheet onde escrevemos
            current_row,                                # linha atual para começar
            segunda_semana,                             # data da segunda-feira desta semana
            semana_idx + 1,                             # índice da semana (1..N)
            escala_semana,                              # dicionário de escala para a semana
            estado,                                     # estado completo (para nomes/cores)
            base_col=2,                                 # começa na coluna B
            mes_restrito=mes_alvo,                      # só mostra dados dos dias deste mês
        )

    nome_ficheiro = SAIDA_XLSX                          # nome do ficheiro de saída (constante definida acima)

    try:                                                # tenta gravar o ficheiro
        wb.save(nome_ficheiro)                          # grava o workbook no disco
        print(f"✅ Planilha gerada: {nome_ficheiro}")    # imprime mensagem de sucesso
    except Exception as e:                              # se der qualquer erro
        print(f"⚠️ Erro ao gravar {nome_ficheiro}: {e}")  # imprime mensagem de erro
